"""
Napkin — File Parser Service

Unified parser for feedback files: CSV, XLSX, JSON, JSONL, TXT, MD, TSV, DOCX, PDF.
Extracts feedback text from any supported format using smart column detection.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# Column names that likely contain feedback text, in priority order.
FEEDBACK_COLUMN_NAMES = [
    "review", "review_body", "review_text",
    "feedback", "feedback_text",
    "comment", "comments",
    "text", "content", "message", "body",
    "description", "summary", "note", "notes",
    "response", "input", "query", "question", "answer",
    "raw_text", "user_feedback", "customer_feedback",
    "opinion", "suggestion", "complaint",
]

# Field names for JSON objects, same priority.
JSON_TEXT_FIELDS = [
    "feedback_text", "feedback", "text", "content", "message",
    "body", "comment", "review", "note", "description",
    "summary", "response", "input", "query", "question", "answer",
]

MAX_ROWS = 5000  # Safety cap


def parse_file(content: bytes, filename: str) -> list[str]:
    """
    Parse a file and return a list of feedback text strings.

    Supports: CSV, XLSX, JSON, JSONL, TXT, MD, TSV, DOCX, PDF.
    Raises ValueError for unsupported formats.
    """
    ext = Path(filename.lower()).suffix

    parsers = {
        ".csv": _parse_csv,
        ".tsv": _parse_tsv,
        ".xlsx": _parse_xlsx,
        ".xls": _parse_xlsx,
        ".json": _parse_json,
        ".jsonl": _parse_jsonl,
        ".txt": _parse_text,
        ".md": _parse_text,
        ".docx": _parse_docx,
        ".pdf": _parse_pdf,
    }

    parser = parsers.get(ext)
    if not parser:
        raise ValueError(f"Unsupported file type: {ext}")

    texts = parser(content)

    # Filter empty strings and cap at MAX_ROWS
    cleaned = [t.strip() for t in texts if t and t.strip()]
    if len(cleaned) > MAX_ROWS:
        logger.warning("file_parser.truncated", filename=filename, original=len(cleaned), kept=MAX_ROWS)
        cleaned = cleaned[:MAX_ROWS]

    logger.info("file_parser.complete", filename=filename, format=ext, texts_extracted=len(cleaned))
    return cleaned


# ============================================================
# CSV — Smart column detection
# ============================================================

def _parse_csv(content: bytes) -> list[str]:
    text = content.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []

    if not fieldnames:
        # Fall back to line-based parsing
        return [line.strip() for line in text.split("\n") if line.strip()]

    col = _find_feedback_column(fieldnames)

    # If we found a column by name, read all rows from it
    if col:
        return [row.get(col, "").strip() for row in reader if (row.get(col) or "").strip()]

    # No column name match — sample rows to find longest-text column
    return _parse_csv_by_longest_column(reader, fieldnames)


def _parse_csv_by_longest_column(reader, fieldnames: list[str]) -> list[str]:
    """Fall back to the column with the longest average text."""
    all_rows = list(reader)
    if not all_rows:
        return []

    sample = all_rows[:50]
    avg_lens = {}
    for fn in fieldnames:
        total = sum(len((r.get(fn) or "")) for r in sample)
        avg_lens[fn] = total / len(sample)

    best_col = max(avg_lens, key=avg_lens.get)  # type: ignore[arg-type]

    # If the best column has very short text (<10 chars avg), concatenate all columns
    if avg_lens[best_col] < 10:
        return [" | ".join((r.get(fn) or "") for fn in fieldnames).strip() for r in all_rows if any(r.values())]

    return [(r.get(best_col) or "").strip() for r in all_rows if (r.get(best_col) or "").strip()]


# ============================================================
# TSV
# ============================================================

def _parse_tsv(content: bytes) -> list[str]:
    text = content.decode("utf-8", errors="ignore")
    lines = text.strip().split("\n")
    if len(lines) < 2:
        return [lines[0]] if lines else []

    headers = lines[0].split("\t")
    col = _find_feedback_column(headers)

    if col:
        col_idx = headers.index(col)
        texts = []
        for line in lines[1:]:
            parts = line.split("\t")
            if col_idx < len(parts) and parts[col_idx].strip():
                texts.append(parts[col_idx].strip())
        return texts

    # Fallback: join all columns
    return [" | ".join(line.split("\t")).strip() for line in lines[1:] if line.strip()]


# ============================================================
# XLSX — Excel spreadsheet
# ============================================================

def _parse_xlsx(content: bytes) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("XLSX support requires openpyxl. Install it with: pip install openpyxl")

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # First row is header
    headers = [str(cell or "").strip() for cell in rows[0]]
    col_idx = None

    # Try to find feedback column by name
    col_name = _find_feedback_column(headers)
    if col_name:
        col_idx = headers.index(col_name)
    else:
        # Find column with longest average text
        sample = rows[1:51]
        if sample:
            avg_lens = {}
            for i, h in enumerate(headers):
                total = sum(len(str(row[i] or "")) for row in sample if i < len(row))
                avg_lens[i] = total / len(sample)
            col_idx = max(avg_lens, key=avg_lens.get)  # type: ignore[arg-type]

    if col_idx is not None:
        texts = []
        for row in rows[1:]:
            if col_idx < len(row):
                val = str(row[col_idx] or "").strip()
                if val:
                    texts.append(val)
        return texts

    # Fallback: concatenate all columns
    return [
        " | ".join(str(cell or "") for cell in row).strip()
        for row in rows[1:]
        if any(cell for cell in row)
    ]


# ============================================================
# JSON
# ============================================================

def _parse_json(content: bytes) -> list[str]:
    text = content.decode("utf-8", errors="ignore")
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        # Try line-by-line as JSONL fallback
        return _parse_jsonl(content)

    items = parsed if isinstance(parsed, list) else [parsed]
    return _extract_json_texts(items)


def _parse_jsonl(content: bytes) -> list[str]:
    text = content.decode("utf-8", errors="ignore")
    items = []
    for line in text.strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        try:
            items.append(json.loads(line))
        except json.JSONDecodeError:
            # Treat as plain text
            items.append(line)

    # Separate strings from dicts
    texts = [item for item in items if isinstance(item, str) and item.strip()]
    dicts = [item for item in items if isinstance(item, dict)]

    if dicts:
        texts.extend(_extract_json_texts(dicts))

    return texts


def _extract_json_texts(items: list[dict]) -> list[str]:
    """Extract text from a list of JSON objects using field priority."""
    texts = []
    for item in items:
        if isinstance(item, str):
            if item.strip():
                texts.append(item.strip())
            continue
        if not isinstance(item, dict):
            continue

        # Try priority fields
        found = False
        for field in JSON_TEXT_FIELDS:
            if field in item and isinstance(item[field], str) and item[field].strip():
                texts.append(item[field].strip())
                found = True
                break

        if not found:
            # Fallback: find the longest string value
            str_vals = [(k, v) for k, v in item.items() if isinstance(v, str) and len(v) > 10]
            if str_vals:
                best = max(str_vals, key=lambda x: len(x[1]))
                texts.append(best[1].strip())
            else:
                # Last resort: dump the whole thing
                dumped = json.dumps(item, default=str)
                if len(dumped) > 5:
                    texts.append(dumped)

    return texts


# ============================================================
# TXT / MD
# ============================================================

def _parse_text(content: bytes) -> list[str]:
    text = content.decode("utf-8", errors="ignore")

    # If text has clear delimiters, split on them
    if "\n---\n" in text:
        return [chunk.strip() for chunk in text.split("\n---\n") if chunk.strip()]
    if "\n\n" in text:
        return [chunk.strip() for chunk in text.split("\n\n") if chunk.strip()]

    # Single lines
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    return lines if lines else [text.strip()] if text.strip() else []


# ============================================================
# DOCX
# ============================================================

def _parse_docx(content: bytes) -> list[str]:
    try:
        from docx import Document
    except ImportError:
        raise ValueError("DOCX support requires python-docx. Install it with: pip install python-docx")

    doc = Document(io.BytesIO(content))
    return [p.text.strip() for p in doc.paragraphs if p.text.strip()]


# ============================================================
# PDF
# ============================================================

def _parse_pdf(content: bytes) -> list[str]:
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ValueError("PDF support requires pypdf. Install it with: pip install pypdf")

    reader = PdfReader(io.BytesIO(content))
    texts = []
    for page in reader.pages:
        text = page.extract_text()
        if text and text.strip():
            texts.append(text.strip())
    return texts


# ============================================================
# Shared: Column name matching
# ============================================================

def _find_feedback_column(headers: list[str]) -> str | None:
    """Find the best feedback column by matching against known names."""
    header_map = {h.lower().strip(): h for h in headers}

    for candidate in FEEDBACK_COLUMN_NAMES:
        if candidate in header_map:
            return header_map[candidate]

    return None
